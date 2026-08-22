from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, text

from fuel_predictor.application.routing import RouteDistance
from fuel_predictor.main import create_app


class RecordingRoutingProvider:
    def __init__(self) -> None:
        self.submitted_sequences: list[tuple[str, ...]] = []

    def calculate_distance(self, stop_sequence: tuple[str, ...]) -> RouteDistance:
        self.submitted_sequences.append(stop_sequence)
        return RouteDistance(total_distance_km=86.4, provider_name="test-routing")


def _train_baseline(client: TestClient) -> None:
    historical_csv = (
        "Kategori ANGBER,Mode Aktivitas,Jam Lifting,Jarak Total (km),"
        "Bahan Bakar Disiapkan (L),Sumber Jarak\n"
        "ANGBER,transport,,20,18,manual\n"
        "ANGBER,transport,,40,28,manual\n"
    )
    dataset = client.post(
        "/api/v1/historical-datasets",
        files={"file": ("riwayat.csv", historical_csv.encode(), "text/csv")},
    ).json()["dataset_version"]
    response = client.post(
        f"/api/v1/dataset-versions/{dataset['dataset_version_id']}/baseline-candidates"
    )
    assert response.status_code == 201
    promoted = client.post(
        f"/api/v1/model-candidates/{response.json()['model_version_id']}/promote"
    )
    assert promoted.status_code == 200


def test_bulk_prediction_templates_are_localized_and_explain_the_columns(tmp_path: Path) -> None:
    with TestClient(create_app(database_path=tmp_path / "operations.sqlite3")) as client:
        excel = client.get("/api/v1/bulk-operation-predictions/template?format=xlsx")
        csv = client.get("/api/v1/bulk-operation-predictions/template?format=csv")

    assert excel.status_code == 200
    assert "attachment" in excel.headers["content-disposition"]
    workbook = load_workbook(BytesIO(excel.content), data_only=True)
    assert workbook.sheetnames == ["Operasi Harian", "Petunjuk"]
    assert workbook["Operasi Harian"]["A1"].value == "Kategori ANGBER (wajib)"
    assert "Kolom wajib" in str(workbook["Petunjuk"]["A1"].value)

    assert csv.status_code == 200
    assert "Kategori ANGBER (wajib)" in csv.text
    assert "Urutan Pemberhentian (opsional)" in csv.text


def test_bulk_prediction_keeps_valid_rows_and_reports_invalid_rows_with_provenance(
    tmp_path: Path,
) -> None:
    provider = RecordingRoutingProvider()
    bulk_csv = (
        "Kategori ANGBER (wajib),Mode Aktivitas (wajib),Jam Lifting (opsional),"
        "Jarak Total (km) (wajib),Sumber Jarak (wajib),Urutan Pemberhentian (opsional)\n"
        "ANGBER,transport,,20,manual,Depo > Site A > Depo\n"
        "ANGBER,lifting,0,25,manual,Depo > Site B > Depo\n"
    )
    with TestClient(
        create_app(database_path=tmp_path / "operations.sqlite3", routing_provider=provider)
    ) as client:
        _train_baseline(client)
        response = client.post(
            "/api/v1/bulk-operation-predictions",
            files={"file": ("rencana-operasi.csv", bulk_csv.encode(), "text/csv")},
        )

    assert response.status_code == 201
    body = response.json()
    assert body["accepted_row_count"] == 1
    assert body["quarantined_row_count"] == 1
    assert body["ignored_blank_row_count"] == 0
    assert provider.submitted_sequences == [("Depo", "Site A", "Depo")]

    accepted = body["accepted_rows"][0]
    assert accepted["source"]["sheet_name"] == "CSV"
    assert accepted["source"]["row_number"] == 2
    assert (
        accepted["source"]["raw_values"]["Urutan Pemberhentian (opsional)"]
        == "Depo > Site A > Depo"
    )
    assert accepted["operation"]["operation_id"].startswith("OPR-")
    assert accepted["operation"]["total_distance_km"] == 86.4
    assert accepted["prediction"]["operation_id"] == accepted["operation"]["operation_id"]
    assert (
        accepted["prediction"]["lineage"]["input_operation_id"]
        == accepted["operation"]["operation_id"]
    )
    assert accepted["prediction"]["estimated_fuel_requirement_liters"] > 0
    assert (
        accepted["prediction"]["recommended_allocation_liters"]
        >= accepted["prediction"]["estimated_fuel_requirement_liters"]
    )

    correction = body["correction_report"][0]
    assert correction["source"]["row_number"] == 3
    assert correction["reasons"] == [
        {
            "field": "lifting_hours",
            "message": "Jam lifting harus lebih besar dari 0 untuk mode yang mencakup lifting.",
        }
    ]

    database_url = f"sqlite+pysqlite:///{(tmp_path / 'operations.sqlite3').as_posix()}"
    with create_engine(database_url).connect() as connection:
        source = connection.execute(
            text(
                "SELECT source_filename, sheet_name, row_number "
                "FROM daily_operation_sources WHERE operation_id = :operation_id"
            ),
            {"operation_id": accepted["operation"]["operation_id"]},
        ).one()
    assert source == ("rencana-operasi.csv", "CSV", 2)


def test_bulk_prediction_requires_a_model_before_creating_any_operation(tmp_path: Path) -> None:
    bulk_csv = (
        "Kategori ANGBER (wajib),Mode Aktivitas (wajib),Jam Lifting (opsional),"
        "Jarak Total (km) (wajib),Sumber Jarak (wajib),Urutan Pemberhentian (opsional)\n"
        "ANGBER,transport,,20,manual,\n"
    )
    database_path = tmp_path / "operations.sqlite3"
    with TestClient(create_app(database_path=database_path)) as client:
        response = client.post(
            "/api/v1/bulk-operation-predictions",
            files={"file": ("rencana-operasi.csv", bulk_csv.encode(), "text/csv")},
        )

    assert response.status_code == 409
    assert response.json()["error"]["code"] == "baseline_model_not_found"
    database_url = f"sqlite+pysqlite:///{database_path.as_posix()}"
    with create_engine(database_url).connect() as connection:
        count = connection.execute(text("SELECT COUNT(*) FROM daily_operations")).scalar_one()
    assert count == 0
