import csv
from datetime import date, datetime, time
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook

from fuel_predictor.application.historical_datasets import (
    HistoricalDatasetImportError,
    SourceSheet,
)
from fuel_predictor.domain.historical_dataset import RawValue


class SpreadsheetHistoricalDatasetSourceReader:
    def read(self, filename: str, content: bytes) -> tuple[SourceSheet, ...]:
        suffix = Path(filename).suffix.casefold()
        if suffix == ".csv":
            return (self._read_csv(content),)
        if suffix == ".xlsx":
            return self._read_xlsx(content)
        raise HistoricalDatasetImportError(
            "Format berkas tidak didukung. Gunakan CSV atau Excel .xlsx."
        )

    def _read_csv(self, content: bytes) -> SourceSheet:
        try:
            decoded = content.decode("utf-8-sig")
        except UnicodeDecodeError as error:
            raise HistoricalDatasetImportError("CSV harus menggunakan pengodean UTF-8.") from error
        try:
            dialect = csv.Sniffer().sniff(decoded, delimiters=",;")
        except csv.Error:
            dialect = csv.excel
        rows = list(csv.reader(StringIO(decoded), dialect=dialect))
        if not rows:
            raise HistoricalDatasetImportError("Berkas impor tidak memiliki baris header.")
        headers = tuple(_header(value) for value in rows[0])
        return SourceSheet(
            name="CSV",
            headers=headers,
            rows=tuple(
                (index, tuple(row[: len(headers)]) + ("",) * (len(headers) - len(row)))
                for index, row in enumerate(rows[1:], start=2)
            ),
        )

    def _read_xlsx(self, content: bytes) -> tuple[SourceSheet, ...]:
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception as error:
            raise HistoricalDatasetImportError("Berkas Excel tidak dapat dibaca.") from error

        sheets: list[SourceSheet] = []
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = tuple(_header(value) for value in rows[0])
            sheets.append(
                SourceSheet(
                    name=worksheet.title,
                    headers=headers,
                    rows=tuple(
                        (
                            index,
                            tuple(_raw_value(value) for value in row)
                            + (None,) * (len(headers) - len(row)),
                        )
                        for index, row in enumerate(rows[1:], start=2)
                    ),
                )
            )
        if not sheets:
            raise HistoricalDatasetImportError("Berkas Excel tidak memiliki lembar data.")
        return tuple(sheets)


def _header(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _raw_value(value: object) -> RawValue:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)
