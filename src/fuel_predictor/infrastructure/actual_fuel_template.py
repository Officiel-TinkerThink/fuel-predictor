import csv
from io import BytesIO, StringIO
from typing import cast

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

BULK_ACTUAL_FUEL_TEMPLATE_HEADERS = (
    "ID Operasi (wajib)",
    "Bahan Bakar Aktual (L) (wajib)",
    "Sumber Pengukuran (opsional)",
)


def csv_template() -> bytes:
    output = StringIO(newline="")
    csv.writer(output).writerow(BULK_ACTUAL_FUEL_TEMPLATE_HEADERS)
    return output.getvalue().encode("utf-8-sig")


def xlsx_template() -> bytes:
    workbook = Workbook()
    worksheet = cast(Worksheet, workbook.active)
    worksheet.title = "Bahan Bakar Aktual"
    worksheet.append(BULK_ACTUAL_FUEL_TEMPLATE_HEADERS)
    _style_header(worksheet)
    for column, width in zip("ABC", (42, 34, 32), strict=True):
        worksheet.column_dimensions[column].width = width
    worksheet.freeze_panes = "A2"

    instructions = workbook.create_sheet("Petunjuk")
    instructions.append(("Kolom", "Status", "Petunjuk"))
    _style_header(instructions)
    instructions.append(("ID Operasi", "Wajib", "Gunakan ID OPR-... yang sudah ada."))
    instructions.append(("Bahan Bakar Aktual (L)", "Wajib", "Masukkan angka lebih besar dari 0."))
    instructions.append(
        (
            "Sumber Pengukuran",
            "Opsional",
            "Gunakan manual_entry, fuel_meter, atau receipt. Kosong berarti spreadsheet_import.",
        )
    )
    instructions.append(
        ("Hasil", "Informasi", "Baris tanpa ID yang cocok atau nilai tidak valid dikarantina.")
    )
    instructions.column_dimensions["A"].width = 30
    instructions.column_dimensions["B"].width = 18
    instructions.column_dimensions["C"].width = 90
    instructions.freeze_panes = "A2"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _style_header(worksheet: Worksheet) -> None:
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="185C43")
