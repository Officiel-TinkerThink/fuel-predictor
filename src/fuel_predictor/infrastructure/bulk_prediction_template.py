import csv
from io import BytesIO, StringIO
from typing import cast

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

BULK_PREDICTION_TEMPLATE_HEADERS = (
    "Kategori ANGBER (wajib)",
    "Kendaraan (opsional)",
    "Mode Aktivitas (wajib)",
    "Jam Lifting (opsional)",
    "Jarak Total (km) (wajib)",
    "Sumber Jarak (wajib)",
    "Urutan Pemberhentian (opsional)",
)


def csv_template() -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(BULK_PREDICTION_TEMPLATE_HEADERS)
    return output.getvalue().encode("utf-8-sig")


def xlsx_template() -> bytes:
    workbook = Workbook()
    worksheet = cast(Worksheet, workbook.active)
    worksheet.title = "Operasi Harian"
    worksheet.append(BULK_PREDICTION_TEMPLATE_HEADERS)
    _style_header(worksheet)
    for column, width in zip("ABCDEFG", (26, 24, 30, 25, 28, 24, 48), strict=True):
        worksheet.column_dimensions[column].width = width
    worksheet.freeze_panes = "A2"

    instructions = workbook.create_sheet("Petunjuk")
    instructions["A1"] = "Kolom wajib dan cara pengisian"
    instructions["A1"].font = Font(bold=True, color="FFFFFF")
    instructions["A1"].fill = PatternFill("solid", fgColor="185C43")
    instructions.append(("Kolom", "Status", "Petunjuk"))
    for cell in instructions[2]:
        cell.font = Font(bold=True)
    instructions.append(("Kategori ANGBER", "Wajib", "Gunakan ANGBER atau Angkutan Berat."))
    instructions.append(
        (
            "Kendaraan",
            "Opsional",
            "Unit yang menjalankan operasi: Prime Mover, Truck Crane 01, Truck Crane 02, "
            "Whellcrane, OFT Tronton, atau OFT Winch Truck. Diisi agar model dapat "
            "membedakan konsumsi tiap kendaraan; baris tanpa kolom ini tetap terbaca.",
        )
    )
    instructions.append(
        ("Mode Aktivitas", "Wajib", "Gunakan transport, lifting, atau transport_and_lifting.")
    )
    instructions.append(("Jam Lifting", "Opsional", "Wajib dan lebih dari 0 untuk mode lifting."))
    instructions.append(("Jarak Total (km)", "Wajib", "Masukkan angka lebih besar dari 0."))
    instructions.append(("Sumber Jarak", "Wajib", "Gunakan manual atau routing_provider."))
    instructions.append(
        (
            "Urutan Pemberhentian",
            "Opsional",
            "Pisahkan lokasi sesuai urutan planner dengan >, misalnya Depo > Site A > Depo.",
        )
    )
    instructions.append(
        (
            "Hasil",
            "Informasi",
            "Setiap baris valid menghasilkan ID operasi, estimasi kebutuhan BBM, "
            "dan alokasi rekomendasi.",
        )
    )
    instructions.column_dimensions["A"].width = 28
    instructions.column_dimensions["B"].width = 16
    instructions.column_dimensions["C"].width = 92
    instructions.freeze_panes = "A3"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _style_header(worksheet: Worksheet) -> None:
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="185C43")
